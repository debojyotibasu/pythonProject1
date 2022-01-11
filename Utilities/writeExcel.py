import openpyxl

workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
sheet = workbook['LoginTest']

sheet.cell(row=1, column=3).value="Age"

# workbook.save("..//excel//testdata.xlsx")

# To update all Age column value at a time
for rows in range(2,7):
    age = rows+23
    for cols in range(3,4):
        sheet.cell(row=rows, column=cols).value = age

workbook.save("..//excel//testdata.xlsx")