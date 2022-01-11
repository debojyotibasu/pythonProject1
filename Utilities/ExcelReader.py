import openpyxl

def getRowCount(path,sheetName):
    # pass
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path,sheetName):
    # pass
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column

def getCellData(path,sheetName,rowNum,colNum):
    # pass
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value

def setCellData(path,sheetName,rowNum,colNum,data):
    # pass
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)

# path = "..//excel//testdata.xlsx"
# sheetName='LoginTest'
# 
# rows = getRowCount(path,sheetName)
# cols = getColCount(path,sheetName)
# 
# print(rows," --- ",cols)
# print(getCellData(path,sheetName,2,1))
# setCellData(path,sheetName,1,4,"DOB")